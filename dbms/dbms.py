# coding=utf-8
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


class DBMS:
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.container = {'id': '',
                          'author': '',
                          'series': '',
                          'title': '',
                          'edition': '',
                          'isbn': '',
                          'publisher': '',
                          'year': '',
                          'pages': '',
                          'size': '',
                          'language': '',
                          'extension': '',
                          'link1': '',
                          'link2': '',
                          'link3': '',
                          'link4': ''
                          }
        self.positions = {'id': 1,
                          'author': 2,
                          'series': 3,
                          'title': 4,
                          'edition': 5,
                          'isbn': 6,
                          'publisher': 7,
                          'year': 8,
                          'pages': 9,
                          'size': 10,
                          'language': 11,
                          'extension': 12,
                          'link1': 13,
                          'link2': 14,
                          'link3': 15,
                          'link4': 16
                          }
        self.activeRow = 2
        for key, value in self.positions.items():
            aCell = self.ws.cell(row=self.activeRow, column=value)
            aCell.value = key
            aCell.font = Font(bold=True)
            aCell.alignment = Alignment(horizontal='center')
        self.activeRow += 1

    def requestContainer(self):
        container = deepcopy(self.container)
        return container

    def dumpContainer(self, container):
        for key, value in container.items():
            if value:
                position = self.positions[key]
                self.ws.cell(row=self.activeRow, column=position).value = value
        self.activeRow += 1

    def save(self, name):
        self.wb.save(name + '.xlsx')
